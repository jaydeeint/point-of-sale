import json
import os
import sys
from datetime import date

import fitz
import openpyxl
from fpdf import FPDF

MM_TO_PT = 72 / 25.4

if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

OUTPUT_DIR = os.path.join(APP_DIR, "bills")
ASSETS_DIR = os.path.join(APP_DIR, "assets")
EXCEL_PATH = os.path.join(APP_DIR, "transactions.xlsx")
INVOICE_COUNTER_PATH = os.path.join(APP_DIR, "invoice_counter.txt")
BUSINESS_INFO_PATH = os.path.join(ASSETS_DIR, "business_info.json")
KEYBINDINGS_PATH = os.path.join(APP_DIR, "keybindings.json")
PREVIEW_PATH = os.path.join(APP_DIR, "preview.pdf")

DEFAULT_KEYBINDINGS = {
    "add_item": "<Control-Return>",
    "generate_bill": "<Control-g>",
    "clear_all": "<Control-Delete>",
}

EXCEL_HEADERS = [
    "Date", "Bill", "Product", "Quantity", "Price", "Subtotal", "Delivery Charge", "Grand Total",
    "Discount", "Tax", "Payment Method", "Invoice #", "Voided", "Voided Date",
]

PAGE_WIDTH = 80
MARGIN = 4
USABLE_WIDTH = PAGE_WIDTH - (2 * MARGIN)

LOGO_EXTENSIONS = (".png", ".jpg", ".jpeg", ".bmp", ".gif")
DEFAULT_BUSINESS_INFO = {"name": "", "address": "", "phone": ""}
PAYMENT_METHODS = ["Cash", "Card", "Other"]


def find_logo_path():
    for ext in LOGO_EXTENSIONS:
        candidate = os.path.join(ASSETS_DIR, f"logo{ext}")
        if os.path.exists(candidate):
            return candidate
    return None


# ---------- Business info ----------

def get_business_info():
    if not os.path.exists(BUSINESS_INFO_PATH):
        return dict(DEFAULT_BUSINESS_INFO)
    try:
        with open(BUSINESS_INFO_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, ValueError):
        return dict(DEFAULT_BUSINESS_INFO)
    info = dict(DEFAULT_BUSINESS_INFO)
    info.update({k: (data.get(k) or "") for k in DEFAULT_BUSINESS_INFO})
    return info


def save_business_info(info):
    os.makedirs(ASSETS_DIR, exist_ok=True)
    payload = {k: (info.get(k) or "").strip() for k in DEFAULT_BUSINESS_INFO}
    with open(BUSINESS_INFO_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


# ---------- Keyboard shortcuts ----------

def get_keybindings():
    if not os.path.exists(KEYBINDINGS_PATH):
        return dict(DEFAULT_KEYBINDINGS)
    try:
        with open(KEYBINDINGS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, ValueError):
        return dict(DEFAULT_KEYBINDINGS)
    bindings = dict(DEFAULT_KEYBINDINGS)
    bindings.update({k: v for k, v in data.items() if k in DEFAULT_KEYBINDINGS and v})
    return bindings


def save_keybindings(bindings):
    payload = {k: bindings.get(k, DEFAULT_KEYBINDINGS[k]) for k in DEFAULT_KEYBINDINGS}
    with open(KEYBINDINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


# ---------- Invoice numbering ----------

def peek_next_invoice_number():
    current = 0
    if os.path.exists(INVOICE_COUNTER_PATH):
        try:
            with open(INVOICE_COUNTER_PATH, "r") as f:
                current = int(f.read().strip() or 0)
        except (ValueError, OSError):
            current = 0
    return current + 1


def get_next_invoice_number():
    next_number = peek_next_invoice_number()
    with open(INVOICE_COUNTER_PATH, "w") as f:
        f.write(str(next_number))
    return next_number


# ---------- Totals (discount + tax) ----------

def parse_discount(raw, items_total):
    raw = (raw or "").strip()
    if not raw:
        return {"amount": 0.0, "is_percent": False, "value": 0.0}

    is_percent = raw.endswith("%")
    number_part = raw[:-1].strip() if is_percent else raw
    value = float(number_part) if number_part else 0.0
    if value < 0:
        raise ValueError("Discount cannot be negative.")

    amount = items_total * (value / 100) if is_percent else value
    amount = max(0.0, min(amount, items_total))
    return {"amount": amount, "is_percent": is_percent, "value": value}


def compute_totals(items, discount_raw, tax_rate, delivery_charge):
    items_total = sum(item["price"] * item["quantity"] for item in items)
    discount = parse_discount(discount_raw, items_total)
    discounted_subtotal = items_total - discount["amount"]
    tax_amount = discounted_subtotal * (tax_rate / 100) if tax_rate else 0.0
    grand_total = discounted_subtotal + tax_amount + delivery_charge

    return {
        "items_total": items_total,
        "discount": discount,
        "discounted_subtotal": discounted_subtotal,
        "tax_rate": tax_rate,
        "tax_amount": tax_amount,
        "delivery_charge": delivery_charge,
        "grand_total": grand_total,
    }


# ---------- PDF generation ----------

def estimate_page_height(items, business_info):
    extra_header_lines = 1 + sum(1 for v in business_info.values() if v)
    header_height = (46 if find_logo_path() else 20) + extra_header_lines * 4
    items_height = len(items) * 11
    footer_height = 50
    safety_buffer = 20
    return header_height + items_height + footer_height + safety_buffer


class BillPDF(FPDF):
    def __init__(self, invoice_number, business_info):
        super().__init__()
        self.invoice_number = invoice_number
        self.business_info = business_info

    def header(self):
        logo_path = find_logo_path()
        if logo_path:
            logo_w = 24
            self.image(logo_path, x=(self.w - logo_w) / 2, y=4, w=logo_w)
            self.set_y(4 + logo_w + 2)
        self.set_font("Helvetica", "B", 12)
        self.cell(0, 6, "RECEIPT", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 7)
        self.cell(0, 4, f"Receipt #: {self.invoice_number:06d}", align="C", new_x="LMARGIN", new_y="NEXT")
        self.cell(0, 4, f"Date: {date.today().isoformat()}", align="C", new_x="LMARGIN", new_y="NEXT")
        for key in ("name", "address", "phone"):
            value = self.business_info.get(key)
            if value:
                self.cell(0, 4, value, align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(2)


def dashed_separator(pdf):
    y = pdf.get_y() + 1
    pdf.set_dash_pattern(dash=1, gap=1)
    pdf.line(MARGIN, y, PAGE_WIDTH - MARGIN, y)
    pdf.set_dash_pattern()
    pdf.set_y(y + 2)


def _totals_line(pdf, label, value_text, bold=False):
    pdf.set_font("Helvetica", "B" if bold else "", 8)
    pdf.cell(USABLE_WIDTH / 2, 5, label)
    pdf.cell(USABLE_WIDTH / 2, 5, value_text, align="R", new_x="LMARGIN", new_y="NEXT")


def build_pdf(items, totals, payment_method, invoice_number, output_path):
    business_info = get_business_info()
    pdf = BillPDF(invoice_number, business_info)
    pdf.set_margins(MARGIN, MARGIN, MARGIN)
    pdf.set_auto_page_break(False)
    page_height = estimate_page_height(items, business_info)
    pdf.add_page(format=(PAGE_WIDTH, page_height))

    for item in items:
        subtotal = item["price"] * item["quantity"]

        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(USABLE_WIDTH, 4, item["name"], new_x="LMARGIN", new_y="NEXT")

        pdf.set_font("Helvetica", "", 8)
        pdf.cell(USABLE_WIDTH / 2, 4, f"{item['quantity']} x {item['price']:.2f}")
        pdf.cell(USABLE_WIDTH / 2, 4, f"{subtotal:.2f}", align="R", new_x="LMARGIN", new_y="NEXT")

        dashed_separator(pdf)

    _totals_line(pdf, "Items Total", f"{totals['items_total']:.2f}")

    discount = totals["discount"]
    if discount["amount"] > 0:
        label = f"Discount ({discount['value']:.0f}%)" if discount["is_percent"] else "Discount"
        _totals_line(pdf, label, f"-{discount['amount']:.2f}")

    if totals["tax_rate"]:
        _totals_line(pdf, f"Tax ({totals['tax_rate']:.2f}%)", f"{totals['tax_amount']:.2f}")

    delivery_display = "None" if totals["delivery_charge"] == 0 else f"{totals['delivery_charge']:.2f}"
    _totals_line(pdf, "Delivery Charge", delivery_display)
    _totals_line(pdf, "Payment", payment_method or "Cash")

    dashed_separator(pdf)
    _totals_line(pdf, "Grand Total", f"{totals['grand_total']:.2f}", bold=True)
    pdf.set_font("Helvetica", "B", 10)

    content_height_mm = min(pdf.get_y() + MARGIN, page_height)

    pdf.output(output_path)
    trim_to_content(output_path, page_height, content_height_mm)


def trim_to_content(pdf_path, page_height_mm, content_height_mm):
    doc = fitz.open(pdf_path)
    page = doc[0]
    width_pt = page.rect.width
    top_pt = page_height_mm * MM_TO_PT
    bottom_pt = (page_height_mm - content_height_mm) * MM_TO_PT
    page.set_mediabox(fitz.Rect(0, bottom_pt, width_pt, top_pt))
    doc.saveIncr()
    doc.close()


# ---------- Excel logging ----------

def _migrate_headers(sheet):
    current = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if len(current) < len(EXCEL_HEADERS):
        for idx in range(len(current) + 1, len(EXCEL_HEADERS) + 1):
            sheet.cell(row=1, column=idx, value=EXCEL_HEADERS[idx - 1])


def _read_rows():
    if not os.path.exists(EXCEL_PATH):
        return []
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(header, row)))
    return rows


def log_transaction(items, totals, payment_method, bill_name, invoice_number):
    if os.path.exists(EXCEL_PATH):
        workbook = openpyxl.load_workbook(EXCEL_PATH)
        sheet = workbook.active
        _migrate_headers(sheet)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Transactions"
        sheet.append(EXCEL_HEADERS)

    bill_date = date.today().isoformat()
    for item in items:
        subtotal = item["price"] * item["quantity"]
        sheet.append([
            bill_date, bill_name, item["name"], item["quantity"], item["price"], subtotal,
            totals["delivery_charge"], totals["grand_total"], totals["discount"]["amount"], totals["tax_amount"],
            payment_method or "Cash", invoice_number, "", "",
        ])

    workbook.save(EXCEL_PATH)


def void_bill(bill_name):
    if not os.path.exists(EXCEL_PATH):
        return False

    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook.active
    _migrate_headers(sheet)
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    try:
        bill_col = header.index("Bill") + 1
        voided_col = header.index("Voided") + 1
        voided_date_col = header.index("Voided Date") + 1
    except ValueError:
        return False

    found = False
    today_str = date.today().isoformat()
    for row_idx in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_idx, column=bill_col).value == bill_name:
            sheet.cell(row=row_idx, column=voided_col, value="VOID")
            sheet.cell(row=row_idx, column=voided_date_col, value=today_str)
            found = True

    if found:
        workbook.save(EXCEL_PATH)
    return found


def get_transactions_list(limit=25):
    rows = _read_rows()
    bills = {}
    for row in rows:
        bill_name = row.get("Bill")
        if not bill_name or bill_name in bills:
            continue
        bills[bill_name] = {
            "bill": bill_name,
            "date": row.get("Date") or "",
            "invoice_number": row.get("Invoice #"),
            "grand_total": row.get("Grand Total") or 0,
            "voided": bool(row.get("Voided")),
        }
    # bills dict preserves insertion order (chronological, since rows are appended in
    # generation order); reverse it for most-recent-first rather than sorting by name,
    # since filenames aren't reliably sortable across the old and new numbering schemes.
    ordered = list(reversed(bills.values()))
    return ordered[:limit]


def get_sales_summary():
    rows = _read_rows()
    bills = {}
    for row in rows:
        bill_name = row.get("Bill")
        if not bill_name or bill_name in bills:
            continue
        bills[bill_name] = {
            "date": row.get("Date") or "",
            "grand_total": row.get("Grand Total") or 0,
            "voided": bool(row.get("Voided")),
        }

    if not bills:
        return None

    today = date.today().isoformat()
    this_month = date.today().strftime("%Y-%m")

    def totals_for(predicate):
        matching = [b["grand_total"] for b in bills.values() if not b["voided"] and predicate(b["date"])]
        return len(matching), sum(matching)

    def voided_for(predicate):
        matching = [b["grand_total"] for b in bills.values() if b["voided"] and predicate(b["date"])]
        return len(matching), sum(matching)

    today_count, today_total = totals_for(lambda d: d == today)
    month_count, month_total = totals_for(lambda d: d.startswith(this_month))
    all_count, all_total = totals_for(lambda d: True)
    voided_count, voided_total = voided_for(lambda d: True)

    return {
        "today": {"date": today, "count": today_count, "total": today_total},
        "month": {"label": this_month, "count": month_count, "total": month_total},
        "all_time": {"count": all_count, "total": all_total},
        "voided": {"count": voided_count, "total": voided_total},
    }


# ---------- Bill generation ----------

def generate_bill(items, discount_raw, tax_rate, delivery_charge, payment_method):
    totals = compute_totals(items, discount_raw, tax_rate, delivery_charge)
    invoice_number = get_next_invoice_number()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"bill_{date.today().isoformat()}_{invoice_number:06d}.pdf"
    output_path = os.path.join(OUTPUT_DIR, filename)

    build_pdf(items, totals, payment_method, invoice_number, output_path)
    log_transaction(items, totals, payment_method, filename, invoice_number)

    return output_path, totals, invoice_number


def build_preview(items, discount_raw, tax_rate, delivery_charge, payment_method):
    totals = compute_totals(items, discount_raw, tax_rate, delivery_charge)
    invoice_number = peek_next_invoice_number()
    build_pdf(items, totals, payment_method, invoice_number, PREVIEW_PATH)
    return PREVIEW_PATH, totals
